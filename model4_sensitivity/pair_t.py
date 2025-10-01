import scipy.stats as stats
import pandas as pd
import numpy as np
from openpyxl import load_workbook

def pair_t(alpha):

    for P in [5]:
        print(f"P: {P}")
        file = f"simulation_result_model4_Queue_fleet20_redeploylimit5_coverage5.3999999999999995.xlsx"
        static = pd.read_excel(file, sheet_name='static')
        ROA = pd.read_excel(file, sheet_name='ROA')
        
        pairt_df = {}
        static_df = {}
        ROA_df = {}

        for column in static.columns:
            # Paired t-test
            statistic, pvalue = stats.ttest_rel(static[column], ROA[column])
            result = "不顯著"
            if pvalue < alpha / 2:
                print(f"{column}: 有顯著差異, p-value {pvalue:.4f}")
                result = "顯著"
            pairt_df[column] = [pvalue, result]

            static_df[column] = [
                static[column].mean(),
                np.sqrt(static[column].var()),
                1.96 * np.sqrt(static[column].var() / len(static[column]))
            ]
            ROA_df[column] = [
                ROA[column].mean(),
                np.sqrt(ROA[column].var()),
                1.96 * np.sqrt(ROA[column].var() / len(ROA[column]))
            ]

        pairt_df = pd.DataFrame(pairt_df)
        static_df = pd.DataFrame(static_df)
        ROA_df = pd.DataFrame(ROA_df)
        all_df = [static_df, ROA_df]

        # === 寫入 pair-t 結果 ===
        with pd.ExcelWriter(file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            pairt_df.to_excel(writer, sheet_name="pair-t(α = 0.05)", index=False)

        # === 附加 static / ROA 統計結果到原始 sheet ===
        book = load_workbook(file)
        all_sheet_name = ['static', 'ROA']

        for i, sheet_name in enumerate(all_sheet_name):
            if sheet_name in book.sheetnames:
                sheet = book[sheet_name]
                start_row = sheet.max_row + 2  # 空一行

                # 使用新的 writer 附加資料
                with pd.ExcelWriter(file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    all_df[i].to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row)

pair_t(alpha=0.05)